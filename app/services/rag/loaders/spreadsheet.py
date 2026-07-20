"""Delimited and workbook extraction with sheet/table structure retained."""
from pathlib import Path
import csv


def extract_csv_text(file_path: str | Path) -> list[dict]:
    with Path(file_path).open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        rows = list(csv.reader(handle))
    return _rows_to_pages(rows, "CSV")


def extract_xlsx_text(file_path: str | Path) -> list[dict]:
    from openpyxl import load_workbook
    workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    pages = []
    for sheet in workbook.worksheets:
        rows = [["" if value is None else str(value) for value in row] for row in sheet.iter_rows(values_only=True)]
        pages.extend(_rows_to_pages(rows, sheet.title))
    workbook.close()
    return pages


def _rows_to_pages(rows: list[list[str]], heading: str) -> list[dict]:
    nonempty = [[cell.strip() for cell in row] for row in rows if any(str(cell).strip() for cell in row)]
    if not nonempty:
        return []
    header = nonempty[0]
    text = "\n".join(" | ".join(header) if index == 0 else " | ".join(row) for index, row in enumerate(nonempty))
    return [{"page": 1, "heading": heading, "text": text}]
